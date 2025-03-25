from pathlib import Path
from typing import Dict, Any, List
import logging
from openpyxl import load_workbook
from processors.base import BaseProcessor

logger = logging.getLogger(__name__)

class ExcelProcessor(BaseProcessor):
    """Processor for Excel documents."""
    
    def __init__(self):
        super().__init__()
        self.supported_extensions = ['.xlsx', '.xls']
    
    async def extract_metadata(self, file_path: Path) -> Dict[str, Any]:
        """Extract metadata from Excel document."""
        try:
            metadata = self._generate_metadata(file_path)
            
            wb = load_workbook(file_path, read_only=True, data_only=True)
            
            metadata.update({
                "sheet_count": len(wb.sheetnames),
                "sheet_names": wb.sheetnames,
                "creator": wb.properties.creator or '',
                "last_modified_by": wb.properties.lastModifiedBy or '',
                "created": wb.properties.created.isoformat() if wb.properties.created else '',
                "modified": wb.properties.modified.isoformat() if wb.properties.modified else '',
                "title": wb.properties.title or '',
                "subject": wb.properties.subject or '',
                "keywords": wb.properties.keywords or ''
            })
            
            wb.close()
            return metadata
            
        except Exception as e:
            logger.error(f"Error extracting Excel metadata: {str(e)}")
            return self._generate_metadata(file_path)
    
    async def extract_text(self, file_path: Path) -> str:
        """Extract text content from Excel document."""
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            text_content = []
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                text_content.append(f"Sheet: {sheet_name}")
                
                for row in sheet.rows:
                    row_text = []
                    for cell in row:
                        if cell.value is not None:  # Changed from if cell.value to handle 0 values
                            row_text.append(str(cell.value))
                    if row_text:
                        text_content.append(" | ".join(row_text))
            
            wb.close()
            return "\n".join(text_content)
            
        except Exception as e:
            logger.error(f"Error extracting Excel text: {str(e)}")
            return ""
    
    async def extract_tables(self, file_path: Path) -> List[Dict[str, Any]]:
        """Extract tables from Excel document."""
        try:
            tables = []
            wb = load_workbook(file_path, read_only=True, data_only=True)
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                
                # Get all rows including headers
                all_rows = []
                for row in sheet.rows:
                    row_data = []
                    for cell in row:
                        # Convert all values to string, handle None and 0 values
                        value = cell.value
                        if value is not None:  # This handles 0 values correctly
                            row_data.append(str(value))
                        else:
                            row_data.append("")
                    if any(cell != "" for cell in row_data):  # Only add non-empty rows
                        all_rows.append(row_data)
                
                if all_rows:  # Only create table if we have data
                    table_data = {
                        "table_id": f"sheet_{sheet_name}",
                        "sheet_name": sheet_name,
                        "rows": all_rows[1:] if len(all_rows) > 1 else [],  # Skip header row for data
                        "headers": all_rows[0] if all_rows else [],
                        "column_count": len(all_rows[0]) if all_rows else 0,
                        "row_count": len(all_rows)
                    }
                    tables.append(table_data)
            
            wb.close()
            return tables
            
        except Exception as e:
            logger.error(f"Error extracting Excel tables: {str(e)}")
            return []
    
    async def extract_charts(self, file_path: Path) -> List[Dict[str, Any]]:
        """Extract charts/graphs from Excel document."""
        try:
            charts = []
            wb = load_workbook(file_path, read_only=False)  # Need read-write mode for charts
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                
                # Get chart information if available
                if hasattr(sheet, '_charts'):
                    for chart in sheet._charts:
                        chart_data = {
                            "chart_id": f"chart_{len(charts) + 1}",
                            "sheet_name": sheet_name,
                            "chart_type": chart.chart_type if hasattr(chart, 'chart_type') else 'unknown',
                            "title": chart.title.text if hasattr(chart, 'title') and chart.title else "Untitled Chart",
                            "position": {
                                "row": chart.anchor._from.row if hasattr(chart, 'anchor') else 0,
                                "col": chart.anchor._from.col if hasattr(chart, 'anchor') else 0
                            }
                        }
                        charts.append(chart_data)
            
            wb.close()
            return charts
            
        except Exception as e:
            logger.error(f"Error extracting Excel charts: {str(e)}")
            return [] 