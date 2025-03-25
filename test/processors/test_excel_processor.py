import asyncio
import os
from pathlib import Path
import pytest
from openpyxl import Workbook
from processors.factory import ProcessorFactory

# Create test data directory if it doesn't exist
TEST_DATA_DIR = Path(__file__).parent / "test_data"
TEST_DATA_DIR.mkdir(exist_ok=True)

def create_test_excel():
    """Create a test Excel file with sample data."""
    wb = Workbook()
    
    # Sheet 1: Sales Data
    ws1 = wb.active
    ws1.title = "Sales"
    
    # Add headers
    headers = ["Date", "Product", "Quantity", "Price", "Total"]
    for col, header in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=header)
    
    # Add data
    data = [
        ("2024-01-01", "Product A", 10, 100, 1000),
        ("2024-01-02", "Product B", 5, 200, 1000),
        ("2024-01-03", "Product C", 8, 150, 1200),
    ]
    
    for row, row_data in enumerate(data, 2):
        for col, value in enumerate(row_data, 1):
            ws1.cell(row=row, column=col, value=value)
    
    # Sheet 2: Summary
    ws2 = wb.create_sheet("Summary")
    ws2.cell(row=1, column=1, value="Total Sales")
    ws2.cell(row=1, column=2, value=3200)
    
    # Save the workbook
    test_file = TEST_DATA_DIR / "test_sales.xlsx"
    wb.save(test_file)
    return test_file

@pytest.mark.asyncio
async def test_excel_processor():
    """Test the Excel processor with a sample file."""
    # Create test Excel file
    test_file = create_test_excel()
    
    try:
        # Get the Excel processor
        processor = ProcessorFactory.get_processor(test_file)
        
        # Process the document
        result = await processor.process(test_file)
        
        # Verify processing status
        assert result["processing_status"] == "success"
        
        # Verify metadata
        metadata = result["metadata"]
        assert metadata["extension"] == ".xlsx"
        assert metadata["sheet_count"] == 2
        assert "Sales" in metadata["sheet_names"]
        assert "Summary" in metadata["sheet_names"]
        
        # Verify content
        content = result["content"]
        assert len(content["text"]) > 0
        assert len(content["tables"]) == 2  # One table per sheet
        
        # Verify Sales table
        sales_table = next(t for t in content["tables"] if t["sheet_name"] == "Sales")
        assert sales_table["column_count"] == 5
        assert sales_table["row_count"] == 4  # Including header row
        assert len(sales_table["headers"]) == 5  # Five columns in header
        assert sales_table["headers"] == ["Date", "Product", "Quantity", "Price", "Total"]
        assert len(sales_table["rows"]) == 3  # Three data rows
        
        # Verify first row of sales data
        first_row = sales_table["rows"][0]
        assert first_row == ["2024-01-01", "Product A", "10", "100", "1000"]
        
        # Verify Summary table
        summary_table = next(t for t in content["tables"] if t["sheet_name"] == "Summary")
        assert summary_table["column_count"] == 2
        assert summary_table["row_count"] == 1  # One row total
        assert len(summary_table["headers"]) == 2  # Two columns
        assert summary_table["headers"] == ["Total Sales", "3200"]  # First row is treated as header
        assert len(summary_table["rows"]) == 0  # No data rows (all content is in header)
        
        print("\nTest Results:")
        print("-" * 50)
        print(f"File: {test_file}")
        print(f"Status: {result['processing_status']}")
        print("\nMetadata:")
        for key, value in metadata.items():
            print(f"  {key}: {value}")
        
        print("\nContent Summary:")
        print(f"  Text Length: {len(content['text'])} characters")
        print(f"  Tables Found: {len(content['tables'])}")
        print(f"  Charts Found: {len(content['charts'])}")
        
        print("\nSales Table:")
        print(f"  Headers: {sales_table['headers']}")
        print(f"  First Row: {sales_table['rows'][0]}")
        
        print("\nSummary Table:")
        print(f"  Headers: {summary_table['headers']}")
        print(f"  Rows: {summary_table['rows']}")
        
    finally:
        # Clean up test file
        if test_file.exists():
            test_file.unlink()

if __name__ == "__main__":
    asyncio.run(test_excel_processor()) 